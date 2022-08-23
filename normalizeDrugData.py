import openpyxl
oldData = openpyxl.load_workbook("D:/Programming/Repositories/irregularVolumeNN/rawData/drug201.xlsx")
normalizedData = openpyxl.load_workbook("D:/Programming/Repositories/irregularVolumeNN/normalization/normalizedDrugData.xlsx")

oldDataWrite = oldData.worksheets[0]
normalizedDataWrite = normalizedData.worksheets[0]

#docs are essentually the same as the other normalization 
def normalizeAge():
    index = 2
    while(oldDataWrite.cell(column=1, row=index).value != None):
        normalizedDataWrite.cell(column=1, row=index).value = (oldDataWrite.cell(column=1, row=index).value - 15) / (74 - 15)
        index += 1

def normalizeNaToK():
    index = 2
    while(oldDataWrite.cell(column=5, row=index).value != None):
        normalizedDataWrite.cell(column=5, row=index).value = (oldDataWrite.cell(column=5, row=index).value - 6.269) / (38.247 - 6.269)
        index += 1
#normalizeAge()
normalizeNaToK()
normalizedData.save("D:/Programming/Repositories/irregularVolumeNN/normalization/normalizedDrugData.xlsx")