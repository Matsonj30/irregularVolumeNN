
import openpyxl


oldData = openpyxl.load_workbook("D:/Programming/Repositories/screenerSettings/highVolume.xlsx", data_only=True) #data only ignores formula aspects 
normalizedData = openpyxl.load_workbook("D:/Programming/Repositories/irregularVolumeNN/normalizedData.xlsx")

oldDataWrite = oldData.worksheets[0]
normalizedDatawrite = normalizedData.worksheets[0]

#(value - 14631) / (209591899- 14631) for each data point, can flip if need to
#(v - min) / (max - min) for each unique column
def normalizeVolume(index):
    while(oldDataWrite.cell(column=4, row=index).value != None):
        normalizedDatawrite.cell(column=4, row=index).value = (oldDataWrite.cell(column=4, row=index).value - 14631) / (209591899- 14631)
        index += 1
        
#use one hot encoding for this
#we have 6 market cap sizes, so we will just have an input node for each market cap
#might just have a cap such that anything over 300M is 1 and then normalize everything in that range since most data is nano-mid cap
def normalizeMktCap(index):
    while(oldDataWrite.cell(column=5, row=index).value != None): 
        mktCap = oldDataWrite.cell(column=5, row=index).value #marketCap in string value
        if(mktCap != '-'):
            slicedCap = int(mktCap[0:-4])
            if(mktCap[-1] == 'M'): #mktCap in millions
                if((slicedCap < 50)): #less than 50 million aka micro cap
                    normalizedDatawrite.cell(column=5, row=index).value = "Nano"
               
                elif(slicedCap < 300):
                    normalizedDatawrite.cell(column=5, row=index).value = "Micro"
                  
                else:
                     normalizedDatawrite.cell(column=5, row=index).value = "Small"
                     
            else: #mktCap in billions
                if(int(slicedCap < 2)):
                    normalizedDatawrite.cell(column=5, row=index).value = "Small"
                  
                elif(int(slicedCap < 10)):
                    normalizedDatawrite.cell(column=5, row=index).value = "Mid"
                 
                elif(int(slicedCap < 200)):
                    normalizedDatawrite.cell(column=5, row=index).value = "Large"
        
                else:
                    normalizedDatawrite.cell(column=5, row=index).value = "Mega"


        else:
             normalizedDatawrite.cell(column=5, row=index).value = "NULL"
        index += 1

#(value - 0.01) / (150.08-0.01) for each data point, can flip if need to
#(v - min) / (max - min) for each unique column
def normalizePrice(index):
    while(oldDataWrite.cell(column=6, row=index).value != None): 
        normalizedDatawrite.cell(column=6, row=index).value = (oldDataWrite.cell(column=6, row=index).value - 0.01) / (150.08-0.01)
        index += 1

#(value - (-0.8286) / (3.7577- (-0.8286)) for each data point, can flip if need to
#(v - min) / (max - min) for each unique column
def normalizeChange(index):
    while(oldDataWrite.cell(column=7, row=index).value != None): 
        normalizedDatawrite.cell(column=7, row=index).value = (oldDataWrite.cell(column=7, row=index).value + 0.8286) / (3.7577 + 0.8286)
        index += 1

def expectedOutPut(index):
    while(oldDataWrite.cell(column=23, row=index).value != None): 
        if(oldDataWrite.cell(column=23, row=index).value == "Green"):
            normalizedDatawrite.cell(column=8, row=index).value = 1
        else:
             normalizedDatawrite.cell(column=8, row=index).value = 0
        index += 1

normalizeVolume(2)
normalizeMktCap(2)
normalizePrice(2)
normalizeChange(2)
expectedOutPut(2)

normalizedData.save("D:/Programming/Repositories/irregularVolumeNN/normalizedData.xlsx")